from os.path import dirname, join
import openpyxl


async def load_excel_file(file_name):
    workbook = openpyxl.load_workbook(join(dirname(__file__), "person.xlsx"))
    sheet = workbook.active
    workbook.close()
    return sheet


async def file_read(data):
    try:
        file_path = join(dirname(__file__), "person.xlsx")
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        for row in data:
            sheet.append(row)
        wb.save(file_path)
        return True
    except Exception as err:
        print(err)


async def check_passport1(name):
    sheet = await load_excel_file("person.xlsx")
    for row in sheet.iter_rows(min_row=2, max_row=1000, values_only=True):
        if row[0] is None:
            break
        if f"{row[0]}" == name:
            return [row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10],
                    row[11], row[12], row[13], row[14]]


async def update_person_info(name, new_data):
    try:
        file_path = join(dirname(__file__), "person.xlsx")
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=1000, values_only=True), start=2):
            if row[0] is None:
                break
            if row[0] == name:
                for i, value in enumerate(new_data[0], start=1):
                    sheet.cell(row=row_num, column=i, value=value)
                wb.save(file_path)
                return True
        return False  # Agar berilgan nomdagi shaxs topilmagan bo'lsa
    except Exception as err:
        print(err)
        return False
